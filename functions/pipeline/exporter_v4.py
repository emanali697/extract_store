"""
v4 exporter - يضيف أعمدة:
- ثقة % (من Vision Judge)
- مصادر متطابقة (Google, OSM, ...)
- نوع التليفون (جوال/أرضي/موحد)
- عدد قراءات التليفون (votes)
- سبب القرار من الـ Judge
"""
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from phone_utils import format_phone_display


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)

ROW_ALT = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

PHONE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
COORD_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
RATING_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

CONF_HIGH = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # ≥80%
CONF_MID = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")   # 50-80
CONF_LOW = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")    # <50

REVIEW_FILL = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


COLUMNS = [
    ("#", 5, "center"),
    ("اسم المتجر", 38, "right"),
    ("الاسم بالإنجليزي", 22, "right"),
    ("التصنيف", 22, "right"),
    ("التليفون", 18, "center"),
    ("نوع الرقم", 10, "center"),
    ("قراءات الرقم", 11, "center"),
    ("الحالة", 10, "center"),
    ("خط العرض", 12, "center"),
    ("خط الطول", 12, "center"),
    ("ثقة المطابقة", 13, "center"),
    ("مصدر المطابقة", 22, "center"),
    ("اسم المتجر (المطابق)", 38, "right"),
    ("العنوان", 40, "right"),
    ("التقييم", 8, "center"),
    ("عدد التقييمات", 12, "center"),
    ("المسافة (م)", 11, "center"),
    ("رابط Google Maps", 40, "right"),
    ("سبب القرار", 38, "right"),
    ("الإطارات", 14, "center"),
    ("ملاحظات", 30, "right"),
    ("محتاج مراجعة", 13, "center"),
    ("سبب المراجعة", 22, "right"),
    ("تاريخ التحليل", 14, "center"),
]


def get_kind_label(kind):
    return {
        'mobile': 'جوال',
        'landline': 'أرضي',
        'unified': 'موحد',
    }.get(kind, '-')


def determine_review(store, v4):
    """قواعد إضافية لـ flag للمراجعة"""
    flags = list(store.get('review_flags') or [])
    conf = (v4 or {}).get('confidence', 0)

    if not v4 or not v4.get('final_match'):
        flags.append('لا توجد مطابقة')
    elif conf < 0.5:
        flags.append(f'ثقة منخفضة ({int(conf*100)}%)')
    elif conf < 0.7:
        flags.append(f'ثقة متوسطة ({int(conf*100)}%)')

    return flags


def build_rows(stores, analysis_date=None):
    if analysis_date is None:
        analysis_date = datetime.now().strftime("%Y-%m-%d")

    rows = []
    for i, s in enumerate(stores, 1):
        v4 = s.get('v4') or {}
        match = v4.get('final_match') or {}

        phone = s.get('phone', '') or ''
        phone_disp = format_phone_display(phone)
        phone_kind = get_kind_label(s.get('phone_kind'))
        phone_votes = s.get('phone_votes', '')

        match_lat = match.get('lat') or s.get('lat')
        match_lng = match.get('lng') or s.get('lng')
        maps_url = match.get('maps_url', '')
        if not maps_url and match_lat and match_lng:
            maps_url = f"https://www.google.com/maps?q={match_lat},{match_lng}"

        sources = match.get('sources', [match.get('_source', '')])
        sources_label = ' + '.join(s.split('_')[0] for s in sources if s)

        flags = determine_review(s, v4)
        needs_review = bool(flags)

        rows.append({
            "#": i,
            "اسم المتجر": s.get('name_ar', ''),
            "الاسم بالإنجليزي": s.get('name_en', ''),
            "التصنيف": s.get('category', ''),
            "التليفون": phone_disp,
            "نوع الرقم": phone_kind,
            "قراءات الرقم": phone_votes,
            "الحالة": s.get('status', 'مفتوح'),
            "خط العرض": match_lat or '',
            "خط الطول": match_lng or '',
            "ثقة المطابقة": f"{int(v4.get('confidence', 0) * 100)}%" if v4 else '0%',
            "مصدر المطابقة": sources_label,
            "اسم المتجر (المطابق)": match.get('name', ''),
            "العنوان": match.get('address', ''),
            "التقييم": match.get('rating', '') or '',
            "عدد التقييمات": match.get('reviews', '') or '',
            "المسافة (م)": match.get('distance_m', '') if match.get('distance_m') is not None else '',
            "رابط Google Maps": maps_url,
            "سبب القرار": v4.get('reasoning', '') if v4 else '',
            "الإطارات": s.get('frame', ''),
            "ملاحظات": s.get('notes', ''),
            "محتاج مراجعة": "نعم" if needs_review else "لا",
            "سبب المراجعة": ", ".join(flags),
            "تاريخ التحليل": analysis_date,
        })
    return rows


def export_excel_v4(stores, output_path, analysis_date=None):
    rows = build_rows(stores, analysis_date)
    df = pd.DataFrame(rows)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "المتاجر"
    ws.sheet_view.rightToLeft = True

    for idx, (name, width, align) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 35
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 28
        base_fill = ROW_ALT if row % 2 == 0 else ROW_WHITE

        conf_str = str(ws.cell(row=row, column=11).value or "0%").replace("%", "")
        try:
            conf_val = int(conf_str)
        except ValueError:
            conf_val = 0
        review_val = str(ws.cell(row=row, column=22).value or "")
        phone_val = str(ws.cell(row=row, column=5).value or "")

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = thin_border
            col_name, _, align = COLUMNS[col - 1]

            if align == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 5 and phone_val and phone_val not in ["", "None", "nan"]:
                cell.fill = PHONE_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="006600")
            elif col in [9, 10]:
                v = str(cell.value or "")
                if v and v not in ["", "None", "nan"]:
                    cell.fill = COORD_FILL
                    cell.font = Font(name="Arial", size=10, color="303F9F")
                else:
                    cell.fill = base_fill
            elif col == 11:  # ثقة
                if conf_val >= 80:
                    cell.fill = CONF_HIGH
                elif conf_val >= 50:
                    cell.fill = CONF_MID
                else:
                    cell.fill = CONF_LOW
                cell.font = Font(name="Arial", size=10, bold=True)
            elif col == 15:  # تقييم
                v = str(cell.value or "")
                if v and v not in ["", "None", "nan"]:
                    cell.fill = RATING_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
                else:
                    cell.fill = base_fill
            elif col == 18:  # رابط
                cell.font = Font(name="Arial", size=9, color="1565C0", underline="single")
            elif col == 22:  # محتاج مراجعة
                if review_val == "نعم":
                    cell.fill = REVIEW_FILL
                    cell.font = Font(name="Arial", size=10, bold=True, color="E65100")
                else:
                    cell.fill = base_fill
            else:
                cell.fill = base_fill

            if col == 2:
                cell.font = Font(name="Arial", size=10, bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return output_path
