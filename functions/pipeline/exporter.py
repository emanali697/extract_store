"""
تصدير النتائج إلى Excel منسق مع أعمدة المراجعة
"""
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)

ROW_ALT = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

PHONE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
COORD_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
RATING_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

MATCH_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
FAR_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
NOTFOUND_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

REVIEW_FILL = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


COLUMNS = [
    ("#", 5, "center"),
    ("اسم المتجر", 40, "right"),
    ("الاسم بالإنجليزي", 22, "right"),
    ("التصنيف", 25, "right"),
    ("رقم الجوال (من الفيديو)", 28, "center"),
    ("رقم الجوال (Places)", 22, "center"),
    ("كل الأرقام", 40, "center"),
    ("الحالة", 10, "center"),
    ("خط العرض", 12, "center"),
    ("خط الطول", 12, "center"),
    ("العنوان", 40, "right"),
    ("التقييم", 8, "center"),
    ("عدد التقييمات", 12, "center"),
    ("المسافة (م)", 10, "center"),
    ("حالة المطابقة", 14, "center"),
    ("رابط Google Maps", 45, "right"),
    ("الإطارات", 14, "center"),
    ("التوقيت", 12, "center"),
    ("ملاحظات", 35, "right"),
    ("محتاج مراجعة", 14, "center"),
    ("سبب المراجعة", 30, "right"),
    ("مراجعة Claude", 40, "right"),
    ("تاريخ التحليل", 14, "center"),
]


def build_rows(stores, analysis_date=None):
    """تحويل stores إلى rows للـ Excel"""
    if analysis_date is None:
        analysis_date = datetime.now().strftime("%Y-%m-%d")

    rows = []
    for i, s in enumerate(stores, 1):
        places = s.get('places', {}) or {}

        video_phone = s.get('phone', '') or ''
        places_phone = places.get('phone', '') or ''
        all_phones = video_phone
        if places_phone and places_phone not in video_phone:
            all_phones = f"{video_phone}, {places_phone}" if video_phone else places_phone

        maps_url = places.get('maps_url', '')
        lat = places.get('lat', '') or s.get('lat', '')
        lng = places.get('lng', '') or s.get('lng', '')
        if not maps_url and lat:
            maps_url = f"https://www.google.com/maps?q={lat},{lng}"

        review_flags = s.get('review_flags', [])
        needs_review = s.get('needs_review', False)

        rows.append({
            "#": i,
            "اسم المتجر": s.get('name_ar', ''),
            "الاسم بالإنجليزي": s.get('name_en', ''),
            "التصنيف": s.get('category', ''),
            "رقم الجوال (من الفيديو)": video_phone,
            "رقم الجوال (Places)": places_phone,
            "كل الأرقام": all_phones,
            "الحالة": s.get('status', 'مفتوح'),
            "خط العرض": lat,
            "خط الطول": lng,
            "العنوان": places.get('address', ''),
            "التقييم": places.get('rating', ''),
            "عدد التقييمات": places.get('reviews', ''),
            "المسافة (م)": places.get('distance_m', ''),
            "حالة المطابقة": places.get('match_status', ''),
            "رابط Google Maps": maps_url,
            "الإطارات": s.get('frame', ''),
            "التوقيت": s.get('timestamp', ''),
            "ملاحظات": s.get('notes', ''),
            "محتاج مراجعة": "نعم" if needs_review else "لا",
            "سبب المراجعة": ", ".join(review_flags),
            "مراجعة Claude": s.get('claude_review', ''),
            "تاريخ التحليل": analysis_date,
        })
    return rows


def export_excel(stores, output_path, analysis_date=None):
    """تصدير Excel منسق"""
    rows = build_rows(stores, analysis_date)
    df = pd.DataFrame(rows)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Styling
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "المتاجر"
    ws.sheet_view.rightToLeft = True

    # أعمدة
    for idx, (name, width, align) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Header
    ws.row_dimensions[1].height = 35
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 28
        base_fill = ROW_ALT if row % 2 == 0 else ROW_WHITE

        phone_val = str(ws.cell(row=row, column=7).value or "")  # كل الأرقام
        match_val = str(ws.cell(row=row, column=15).value or "")  # حالة المطابقة
        review_val = str(ws.cell(row=row, column=20).value or "")  # محتاج مراجعة

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = thin_border
            col_name, _, align = COLUMNS[col - 1]

            if align == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # أرقام الهواتف
            if col in [5, 6, 7] and phone_val and phone_val not in ["", "None", "nan"]:
                cell.fill = PHONE_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="006600")
            # الإحداثيات
            elif col in [9, 10]:
                v = str(cell.value or "")
                if v and v not in ["", "None", "nan"]:
                    cell.fill = COORD_FILL
                    cell.font = Font(name="Arial", size=10, color="303F9F")
                else:
                    cell.fill = base_fill
            # التقييم
            elif col == 12:
                v = str(cell.value or "")
                if v and v not in ["", "None", "nan"]:
                    cell.fill = RATING_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
                else:
                    cell.fill = base_fill
            # حالة المطابقة
            elif col == 15:
                if match_val == 'مطابق':
                    cell.fill = MATCH_FILL
                elif match_val == 'بعيد':
                    cell.fill = FAR_FILL
                elif match_val == 'غير موجود':
                    cell.fill = NOTFOUND_FILL
                cell.font = Font(name="Arial", size=10, bold=True)
            # Maps URL
            elif col == 16:
                cell.font = Font(name="Arial", size=9, color="1565C0", underline="single")
            # محتاج مراجعة
            elif col == 20:
                if review_val == "نعم":
                    cell.fill = REVIEW_FILL
                    cell.font = Font(name="Arial", size=10, bold=True, color="E65100")
                else:
                    cell.fill = base_fill
            else:
                cell.fill = base_fill

            # اسم المتجر bold
            if col == 2:
                cell.font = Font(name="Arial", size=10, bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return output_path
