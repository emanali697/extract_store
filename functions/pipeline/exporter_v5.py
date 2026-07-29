"""
v5 exporter - أعمدة واضحة:
- مصدر الموقع (Google / فيديو) + لون مختلف
- الإحداثيات النهائية (اللي المفروض تستخدمها)
- المسافة من GPS الفيديو (لو مطابقة Google)
- ثقة المطابقة + سبب
"""
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from phone_utils import format_phone_display


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
ROW_ALT = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

PHONE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
GOOGLE_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
FRAME_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
RATING_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COLUMNS = [
    ("#", 5, "center"),
    ("اسم المتجر", 36, "right"),
    ("التصنيف", 20, "right"),
    ("التليفون", 16, "center"),
    ("الحالة", 9, "center"),
    ("خط العرض", 12, "center"),
    ("خط الطول", 12, "center"),
    ("مصدر الموقع", 18, "center"),
    ("دقة الموقع", 14, "center"),
    ("الاسم في Google", 30, "right"),
    ("العنوان", 36, "right"),
    ("التقييم", 8, "center"),
    ("عدد التقييمات", 12, "center"),
    ("المسافة من الفيديو (م)", 18, "center"),
    ("رابط Google Maps", 38, "right"),
    ("سبب المطابقة", 36, "right"),
    ("الإطارات", 14, "center"),
    ("ملاحظات OCR", 28, "right"),
    ("محتاج مراجعة", 13, "center"),
    ("تاريخ التحليل", 13, "center"),
]


STATUS_LABEL = {
    'confirmed_high': 'Google ✅',
    'confirmed_medium': 'Google 🟡',
    'frame_only': 'فيديو 📍',
}

ACCURACY_LABEL = {
    'confirmed_high': '< 50م (مؤكد)',
    'confirmed_medium': '< 50م (متوسط)',
    'frame_only': '~15م ± (تقريبي)',
}


def build_rows(stores, analysis_date=None):
    if analysis_date is None:
        analysis_date = datetime.now().strftime("%Y-%m-%d")

    rows = []
    for i, s in enumerate(stores, 1):
        v5 = s.get('v5') or {}
        cand = v5.get('candidate') or {}
        status = v5.get('status', 'frame_only')

        # الإحداثيات النهائية
        if status == 'frame_only':
            final_lat, final_lng = s.get('lat', ''), s.get('lng', '')
            maps_url = f"https://www.google.com/maps?q={final_lat},{final_lng}" if final_lat else ''
        else:
            final_lat = cand.get('lat', '')
            final_lng = cand.get('lng', '')
            maps_url = cand.get('maps_url') or (f"https://www.google.com/maps?q={final_lat},{final_lng}" if final_lat else '')

        phone = s.get('phone', '') or ''
        phone_disp = format_phone_display(phone)

        flags = []
        if status == 'frame_only':
            flags.append('غير ممسوح في Google')
        elif status == 'confirmed_medium':
            flags.append('مطابقة متوسطة')

        rows.append({
            "#": i,
            "اسم المتجر": s.get('name_ar', ''),
            "التصنيف": s.get('category', ''),
            "التليفون": phone_disp,
            "الحالة": s.get('status', 'مفتوح'),
            "خط العرض": final_lat,
            "خط الطول": final_lng,
            "مصدر الموقع": STATUS_LABEL.get(status, '-'),
            "دقة الموقع": ACCURACY_LABEL.get(status, '-'),
            "الاسم في Google": cand.get('name', '') if status != 'frame_only' else '',
            "العنوان": cand.get('address', '') if status != 'frame_only' else '',
            "التقييم": cand.get('rating', '') if status != 'frame_only' else '',
            "عدد التقييمات": cand.get('reviews', '') if status != 'frame_only' else '',
            "المسافة من الفيديو (م)": cand.get('distance_m', '') if status != 'frame_only' else '',
            "رابط Google Maps": maps_url,
            "سبب المطابقة": v5.get('reason', ''),
            "الإطارات": s.get('frame', ''),
            "ملاحظات OCR": s.get('notes', ''),
            "محتاج مراجعة": "نعم" if flags else "لا",
            "تاريخ التحليل": analysis_date,
        })
    return rows


def export_excel_v5(stores, output_path, analysis_date=None):
    rows = build_rows(stores, analysis_date)
    df = pd.DataFrame(rows)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "المتاجر"
    ws.sheet_view.rightToLeft = True

    for idx, (name, width, align) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 35
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 28
        base_fill = ROW_ALT if row % 2 == 0 else ROW_WHITE
        source_val = str(ws.cell(row=row, column=8).value or "")
        review_val = str(ws.cell(row=row, column=19).value or "")
        phone_val = str(ws.cell(row=row, column=4).value or "")

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = thin_border
            col_name, _, align = COLUMNS[col - 1]

            if align == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 4 and phone_val and phone_val not in ["", "None", "nan"]:
                cell.fill = PHONE_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="006600")
            elif col in [6, 7, 8, 9]:  # موقع + مصدر + دقة
                if "Google" in source_val:
                    cell.fill = GOOGLE_FILL
                elif "فيديو" in source_val:
                    cell.fill = FRAME_FILL
                else:
                    cell.fill = base_fill
                if col == 8:
                    cell.font = Font(name="Arial", size=10, bold=True)
            elif col == 12:  # تقييم
                v = str(cell.value or "")
                if v and v not in ["", "None", "nan"]:
                    cell.fill = RATING_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
                else:
                    cell.fill = base_fill
            elif col == 15:  # رابط
                cell.font = Font(name="Arial", size=9, color="1565C0", underline="single")
                cell.fill = base_fill
            elif col == 19:  # محتاج مراجعة
                if review_val == "نعم":
                    cell.fill = REVIEW_FILL
                    cell.font = Font(name="Arial", size=10, bold=True, color="E65100")
                else:
                    cell.fill = base_fill
            else:
                cell.fill = base_fill

            if col == 2:
                cell.font = Font(name="Arial", size=10, bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return output_path
