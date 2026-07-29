"""
v6 exporter - Excel نهائي مع الحالة (نشط/مقفول/غير محدد) + المصدر.

الأعمدة:
- # | اسم المتجر | التصنيف | التليفون
- خط العرض | خط الطول | مصدر الموقع | دقة الموقع
- الحالة | مصدر الحالة | التفاصيل
- التقييم | عدد التقييمات | ساعات العمل (لو من Google)
- الاسم في Google | العنوان | رابط Maps
- الإطارات | محتاج مراجعة | تاريخ التحليل
"""
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from phone_utils import format_phone_display


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10)
ROW_ALT = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# ألوان الحالة
ACTIVE_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # أخضر
TEMP_CLOSED_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # برتقالي
PERM_CLOSED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # أحمر
UNKNOWN_FILL = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")  # رمادي

# ألوان مصدر الموقع
GOOGLE_LOC_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
FRAME_LOC_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

PHONE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RATING_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COLUMNS = [
    ("#", 5, "center"),
    ("اسم المتجر", 32, "right"),
    ("التصنيف", 18, "right"),
    ("التليفون", 18, "center"),
    ("خط العرض", 11, "center"),
    ("خط الطول", 11, "center"),
    ("مصدر الموقع", 14, "center"),
    ("دقة الموقع", 16, "center"),
    ("الحالة", 11, "center"),
    ("مصدر الحالة", 22, "center"),
    ("تفاصيل الحالة", 42, "right"),
    ("التقييم", 8, "center"),
    ("عدد التقييمات", 12, "center"),
    ("الاسم في Google", 30, "right"),
    ("العنوان", 36, "right"),
    ("رابط Google Maps", 38, "right"),
    ("الإطارات في الفيديو", 16, "center"),
    ("ملاحظات", 28, "right"),
    ("تاريخ التحليل", 12, "center"),
]


LOC_SOURCE_LABEL = {
    'confirmed_high': 'Google ✅',
    'confirmed_medium': 'Google 🟡',
    'frame_only': 'فيديو 📍',
}

LOC_ACCURACY = {
    'confirmed_high': '< 50م (مؤكد)',
    'confirmed_medium': '< 50م (متوسط)',
    'frame_only': '~15م ± (تقريبي)',
}


def build_rows(stores):
    today = datetime.now().strftime("%Y-%m-%d")
    rows = []
    for i, s in enumerate(stores, 1):
        v5 = s.get('v5') or {}
        cand = v5.get('candidate') or {}
        loc_status = v5.get('status', 'frame_only')
        sc = s.get('status_check') or {}

        # الإحداثيات
        if loc_status == 'frame_only':
            lat, lng = s.get('lat', ''), s.get('lng', '')
            maps_url = f"https://www.google.com/maps?q={lat},{lng}" if lat else ''
        else:
            lat = cand.get('lat', '')
            lng = cand.get('lng', '')
            maps_url = cand.get('maps_url') or (f"https://www.google.com/maps?q={lat},{lng}" if lat else '')

        # تفاصيل من Tier 1
        rating = sc.get('rating', '')
        review_count = sc.get('review_count', '')

        rows.append({
            "#": i,
            "اسم المتجر": s.get('name_ar', ''),
            "التصنيف": s.get('category', ''),
            "التليفون": format_phone_display(s.get('phone', '') or ''),
            "خط العرض": lat,
            "خط الطول": lng,
            "مصدر الموقع": LOC_SOURCE_LABEL.get(loc_status, '-'),
            "دقة الموقع": LOC_ACCURACY.get(loc_status, '-'),
            "الحالة": sc.get('status', 'غير محدد'),
            "مصدر الحالة": sc.get('source', '-'),
            "تفاصيل الحالة": sc.get('evidence', ''),
            "التقييم": rating,
            "عدد التقييمات": review_count,
            "الاسم في Google": cand.get('name', '') if loc_status != 'frame_only' else '',
            "العنوان": cand.get('address', '') if loc_status != 'frame_only' else '',
            "رابط Google Maps": maps_url,
            "الإطارات في الفيديو": s.get('frame', ''),
            "ملاحظات": s.get('notes', ''),
            "تاريخ التحليل": today,
        })
    return rows


def export_excel_v6(stores, output_path):
    rows = build_rows(stores)
    df = pd.DataFrame(rows)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "المتاجر"
    ws.sheet_view.rightToLeft = True

    for idx, (name, width, align) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 35
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 32
        base_fill = ROW_ALT if row % 2 == 0 else ROW_WHITE
        loc_source_val = str(ws.cell(row=row, column=7).value or "")
        status_val = str(ws.cell(row=row, column=9).value or "")
        phone_val = str(ws.cell(row=row, column=4).value or "")

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = thin_border
            col_name, _, align = COLUMNS[col - 1]

            if align == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 4 and phone_val and phone_val not in ["", "None", "nan"]:
                cell.fill = PHONE_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="006600")
            elif col in [5, 6, 7, 8]:  # موقع + مصدر + دقة
                if "Google" in loc_source_val:
                    cell.fill = GOOGLE_LOC_FILL
                elif "فيديو" in loc_source_val:
                    cell.fill = FRAME_LOC_FILL
                else:
                    cell.fill = base_fill
                if col == 7:
                    cell.font = Font(name="Arial", size=10, bold=True)
            elif col == 9:  # الحالة
                if status_val == 'نشط':
                    cell.fill = ACTIVE_FILL
                    cell.font = Font(name="Arial", size=10, bold=True, color="1B5E20")
                elif status_val == 'مقفول مؤقت':
                    cell.fill = TEMP_CLOSED_FILL
                    cell.font = Font(name="Arial", size=10, bold=True, color="E65100")
                elif status_val == 'مقفول دائم':
                    cell.fill = PERM_CLOSED_FILL
                    cell.font = Font(name="Arial", size=10, bold=True, color="B71C1C")
                else:
                    cell.fill = UNKNOWN_FILL
                    cell.font = Font(name="Arial", size=10, color="616161")
            elif col == 12:  # تقييم
                v = str(cell.value or "")
                if v and v not in ["", "None", "nan"]:
                    cell.fill = RATING_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
                else:
                    cell.fill = base_fill
            elif col == 16:  # رابط
                cell.font = Font(name="Arial", size=9, color="1565C0", underline="single")
                cell.fill = base_fill
            else:
                cell.fill = base_fill

            if col == 2:
                cell.font = Font(name="Arial", size=10, bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    import json, sys
    sys.stdout.reconfigure(encoding='utf-8')
    in_path = r'd:/sharea elnassim/الشارع الجديد/output_v6/stores_v6_final.json'
    out_path = r'd:/sharea elnassim/الشارع الجديد/output_v6/stores_v6_final.xlsx'
    with open(in_path, 'r', encoding='utf-8') as f:
        stores = json.load(f)
    export_excel_v6(stores, out_path)
    print(f"✅ {out_path}")
