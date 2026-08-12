"""Create a styled Excel workbook from canonical reviewed results."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "#", "اسم المتجر", "التصنيف", "التليفون", "خط العرض", "خط الطول",
    "مصدر الموقع", "دقة الموقع", "الحالة", "مصدر النتيجة", "تفاصيل الحالة",
    "التقييم", "عدد التقييمات", "الاسم بالإنجليزية", "الشارع", "المدينة",
    "الحي", "رابط Google Maps", "حالة المراجعة",
]
WIDTHS = [6, 36, 24, 18, 14, 14, 20, 16, 14, 24, 48, 11, 15, 30, 24, 18, 18, 44, 22]
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ALT_FILL = PatternFill("solid", fgColor="F3F6F9")
PHONE_FILL = PatternFill("solid", fgColor="E8F5E9")
BORDER = Border(**{
    side: Side(style="thin", color="D9E1F2")
    for side in ("left", "right", "top", "bottom")
})


def _maps_url(store: dict[str, Any]) -> str:
    lat, lng = store.get("lat"), store.get("lng")
    return "" if lat is None or lng is None else f"https://www.google.com/maps?q={lat},{lng}"


def _review_label(store: dict[str, Any]) -> str:
    if store.get("review_status") == "approved" or store.get("approved"):
        return "تمت المراجعة والموافقة"
    if store.get("auto_review_decision") == "auto_passed":
        return "مراجعة آلية - مقبول"
    if store.get("auto_review_decision") == "needs_human":
        return "يحتاج مراجعة يدوية"
    return ""


def write_results_excel(
    stores: list[dict[str, Any]],
    output_path: str | Path,
    *,
    job_id: str = "",
) -> Path:
    """Write the current post-review store list to a fresh XLSX file."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "المتاجر بعد المراجعة"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.append(HEADERS)
    for index, store in enumerate(stores, 1):
        ws.append([
            store.get("id", index), store.get("name") or store.get("name_ar") or "",
            store.get("category", ""), str(store.get("phone") or ""), store.get("lat", ""),
            store.get("lng", ""), store.get("location_source", ""),
            store.get("distance") or store.get("location_accuracy_m") or "", store.get("status", ""),
            store.get("source", ""), store.get("evidence", ""), store.get("rating", ""),
            store.get("review_count", ""), store.get("name_en", ""), store.get("street", ""),
            store.get("city", ""), store.get("district", ""), _maps_url(store), _review_label(store),
        ])
    for column, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    for cell in ws[1]:
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 34
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 30
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=column)
            cell.font = Font(name="Arial", size=10, bold=column == 2)
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 4, 5, 6, 8, 9, 12, 13, 19} else "right",
                vertical="center", wrap_text=True,
            )
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL
        phone_cell = ws.cell(row=row, column=4)
        phone_cell.number_format = "@"
        if phone_cell.value:
            phone_cell.fill = PHONE_FILL
        maps_cell = ws.cell(row=row, column=18)
        if maps_cell.value:
            maps_cell.hyperlink = maps_cell.value
            maps_cell.style = "Hyperlink"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("ملخص")
    summary.sheet_view.rightToLeft = True
    summary.append(["البيان", "القيمة"])
    summary.append(["رقم المهمة", job_id])
    summary.append(["عدد المتاجر بعد المراجعة", len(stores)])
    summary.append(["وقت إنشاء الملف", datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S")])
    for cell in summary[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 28
    wb.save(output_path)
    return output_path
